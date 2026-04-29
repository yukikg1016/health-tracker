"""
Health Tracker - Screenshot to Excel
起動: streamlit run health_tracker_app.py
"""

import datetime
import os
import pathlib
import sys
import tempfile

import streamlit as st
from dotenv import load_dotenv

BASE_DIR = pathlib.Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))

# ── カスタムモジュールのキャッシュを毎回クリア（デプロイ後の古いコード防止）──
for _mn in [k for k in list(sys.modules.keys())
            if k.startswith(("excel_writer", "extractors", "gdrive_helper", "autosleep",
                              "dashboard_reader", "dashboard_ui"))]:
    del sys.modules[_mn]

# .env 読み込み
load_dotenv(BASE_DIR / ".env")

# ── クッキー初期化（モジュールレベルで必須）────────────────────────────────
try:
    from streamlit_cookies_controller import CookieController
    _cookie = CookieController()
except Exception:
    _cookie = None

# ── ユーザー設定 ──────────────────────────────────────────────────────────────
def _get_user_passwords() -> dict:
    """Returns {user_id: password} from Streamlit secrets or env vars."""
    users = {}
    # Streamlit Secrets から読む（優先）
    try:
        import streamlit as _st
        for uid in ["1", "2", "3", "4"]:
            key = f"USER_{uid}_PASSWORD"
            try:
                users[uid] = _st.secrets[key]
            except Exception:
                pass
    except Exception:
        pass
    # 環境変数からも読む（ローカル用）
    for uid in ["1", "2", "3", "4"]:
        key = f"USER_{uid}_PASSWORD"
        val = os.environ.get(key)
        if val and uid not in users:
            users[uid] = val
    # フォールバック: 既存の APP_PASSWORD を user 1 に割り当て
    if "1" not in users:
        users["1"] = os.environ.get("APP_PASSWORD", "yukihealth2026")
    return users

def _user_display_name(uid: str) -> str:
    try:
        import streamlit as _st
        return _st.secrets.get(f"USER_{uid}_NAME", f"User {uid}")
    except Exception:
        return os.environ.get(f"USER_{uid}_NAME", f"User {uid}")

# ── パスワード保護 ────────────────────────────────────────────────────────────
def _check_password() -> bool:
    # クッキーで永続ログイン確認
    if _cookie is not None:
        try:
            saved = _cookie.get("ht_user")
            if saved and saved in _get_user_passwords():
                st.session_state["authenticated"] = True
                st.session_state["user_id"] = saved
        except Exception:
            pass

    if st.session_state.get("authenticated"):
        return True

    st.markdown("## 🔐 Health Tracker")
    uid = st.text_input("ユーザーID（1〜4）", key="uid_input", placeholder="1")
    pw  = st.text_input("パスワード", type="password", key="pw_input")
    if st.button("ログイン", type="primary"):
        users = _get_user_passwords()
        if uid in users and pw == users[uid]:
            st.session_state["authenticated"] = True
            st.session_state["user_id"] = uid
            if _cookie is not None:
                try:
                    _cookie.set("ht_user", uid, max_age=30 * 24 * 60 * 60)
                except Exception:
                    pass
            st.rerun()
        else:
            st.error("IDまたはパスワードが違います")
    return False

if not _check_password():
    st.stop()

# ── ユーザーコンテキスト ──────────────────────────────────────────────────────
_USER_ID     = st.session_state.get("user_id", "1")
# User 1 は既存シートをそのまま使う（プレフィックスなし）
_USER_PREFIX = "" if _USER_ID == "1" else f"{_USER_ID}_"

# クラウド環境かローカルかを判定
_IS_CLOUD = not (pathlib.Path.home() / "　Reasearch" / "Yuki データ.xlsx").exists()
DEFAULT_EXCEL = str(pathlib.Path.home() / "　Reasearch" / "Yuki データ.xlsx")

SHEET_OPTIONS = {
    "🍽️ Nutrition（食事・栄養）":    "nutrition",
    "😴 Sleep（睡眠）":              "sleep",
    "🩸 Labs（血液検査）":            "labs",
    "⚖️ In Body（体組成）":          "inbody",
    "💪 Workout":                    "workout",
    "🏋️ Performance（体力測定）":    "performance",
}

_PORT = os.environ.get("PORT", "8501")
st.set_page_config(page_title="Health Tracker", page_icon="🏃", layout="wide")

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Score card */
.score-card {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border-radius: 16px;
    padding: 24px;
    text-align: center;
    color: white;
    box-shadow: 0 8px 32px rgba(0,0,0,0.3);
}
.score-number {
    font-size: 72px;
    font-weight: 900;
    line-height: 1;
    margin: 8px 0;
}
.score-label {
    font-size: 13px;
    letter-spacing: 3px;
    text-transform: uppercase;
    opacity: 0.7;
}
.score-grade {
    font-size: 18px;
    font-weight: 600;
    margin-top: 8px;
}
/* Training intensity badge */
.intensity-badge {
    border-radius: 12px;
    padding: 20px;
    text-align: center;
    font-weight: 700;
    font-size: 22px;
    letter-spacing: 1px;
    color: white;
    box-shadow: 0 4px 15px rgba(0,0,0,0.2);
}
/* Bio age card */
.bio-age-card {
    background: linear-gradient(135deg, #2d1b69 0%, #11998e 100%);
    border-radius: 16px;
    padding: 20px;
    color: white;
    box-shadow: 0 8px 32px rgba(0,0,0,0.3);
}
/* Advice card */
.advice-card {
    background: #1e1e2e;
    border-left: 4px solid #7c3aed;
    border-radius: 8px;
    padding: 14px 18px;
    color: #e2e8f0;
    margin: 6px 0;
    font-size: 14px;
    line-height: 1.6;
}
/* Component bar */
.comp-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 6px 0;
    border-bottom: 1px solid rgba(255,255,255,0.08);
    color: #cbd5e1;
    font-size: 13px;
}
</style>
""", unsafe_allow_html=True)

st.title("🏃 Health Tracker")

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 設定")
    provider = st.radio("AI Provider", ["gemini", "claude"], horizontal=True, key="provider")

    env_key = os.environ.get("GEMINI_API_KEY" if provider == "gemini" else "ANTHROPIC_API_KEY", "")
    if "api_key_input" not in st.session_state:
        st.session_state["api_key_input"] = env_key

    api_key = st.text_input(
        "APIキー", type="password", key="api_key_input",
        help="環境変数 GEMINI_API_KEY / ANTHROPIC_API_KEY でも設定可能",
    )

    # クラウドモード：Google Drive連携
    if _IS_CLOUD:
        from gdrive_helper import is_configured
        if is_configured():
            st.caption("☁️ Google Drive連携 有効")
        else:
            st.caption("⚠️ Google Drive未設定（SecretsにGDRIVE_FILE_IDとGOOGLE_SERVICE_ACCOUNT_JSONが必要）")
        excel_path = None  # クラウドではパス不要
    else:
        excel_path = st.text_input("Excelファイルパス", value=DEFAULT_EXCEL, key="excel_path_input")

    st.divider()
    st.caption(f"👤 ログイン中: {_user_display_name(_USER_ID)}（ID: {_USER_ID}）")
    if st.button("🔓 ログアウト", use_container_width=True):
        st.session_state.clear()
        if _cookie is not None:
            try:
                _cookie.remove("ht_user")
            except Exception:
                pass
        st.rerun()
    st.caption("使い方: 画像をアップロード → 抽出 → 確認 → 書き込み")

# ── Main tabs ────────────────────────────────────────────────────────────────
tab_dash, tab_input = st.tabs(["📊 ダッシュボード", "📸 データ入力"])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
with tab_dash:
    if not api_key:
        st.info("サイドバーでAPIキーを入力するとダッシュボードが表示されます")
    else:
        # Load Excel data
        dash_excel_path = None
        if _IS_CLOUD:
            if "dashboard_excel_path" not in st.session_state:
                with st.spinner("データを読み込み中..."):
                    try:
                        from gdrive_helper import download_to_temp
                        st.session_state["dashboard_excel_path"] = download_to_temp()
                    except Exception as e:
                        st.error(f"データ取得エラー: {e}")
            dash_excel_path = st.session_state.get("dashboard_excel_path")
        else:
            dash_excel_path = excel_path if excel_path and pathlib.Path(excel_path).exists() else None

        if not dash_excel_path:
            st.warning("Excelファイルが見つかりません")
        else:
            # Refresh button
            rcol1, rcol2 = st.columns([4, 1])
            with rcol2:
                if st.button("🔄 更新", use_container_width=True):
                    st.session_state.pop("dashboard_excel_path", None)
                    st.session_state.pop("dashboard_data", None)
                    st.rerun()

            # Cache dashboard data in session
            if "dashboard_data" not in st.session_state:
                with st.spinner("データを分析中..."):
                    try:
                        from dashboard_reader import read_dashboard_data
                        st.session_state["dashboard_data"] = read_dashboard_data(
                            dash_excel_path, sheet_prefix=_USER_PREFIX)
                    except Exception as e:
                        st.error(f"データ読み込みエラー: {e}")
                        st.session_state["dashboard_data"] = {}

            dd = st.session_state.get("dashboard_data", {})
            sleep_recs   = dd.get("sleep", [])
            workout_recs = dd.get("workout", [])
            nutrition_recs = dd.get("nutrition", [])
            inbody = dd.get("inbody")
            labs   = dd.get("labs")

            from dashboard_ui import (calc_recovery_score, calc_training_recommendation,
                                       calc_biological_age, build_ai_prompt, get_ai_advice)

            recovery = calc_recovery_score(sleep_recs)
            training = calc_training_recommendation(recovery, workout_recs)
            bio_age  = calc_biological_age(inbody, sleep_recs, labs, workout_recs)

            today_str = datetime.date.today().strftime("%Y年%m月%d日")
            st.markdown(f"<p style='color:#94a3b8;font-size:13px;margin-bottom:16px'>"
                        f"最終更新: {today_str}</p>", unsafe_allow_html=True)

            # ── Row 1: Recovery Score + Training Intensity + Bio Age ──────────
            r1c1, r1c2, r1c3 = st.columns([1, 1, 1])

            with r1c1:
                score = recovery.get("score")
                if score is not None:
                    grade_map = {
                        "excellent": ("⚡ EXCELLENT", "#00ff88"),
                        "good":      ("✅ GOOD",      "#4ade80"),
                        "moderate":  ("⚠️ MODERATE",  "#fbbf24"),
                        "poor":      ("🔴 POOR",      "#f87171"),
                    }
                    grade_label, grade_color = grade_map.get(recovery["grade"], ("—", "#94a3b8"))
                    st.markdown(f"""
                    <div class="score-card">
                        <div class="score-label">RECOVERY SCORE</div>
                        <div class="score-number" style="color:{grade_color}">{score}</div>
                        <div class="score-grade" style="color:{grade_color}">{grade_label}</div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class="score-card">
                        <div class="score-label">RECOVERY SCORE</div>
                        <div class="score-number" style="color:#475569">—</div>
                        <div class="score-grade" style="color:#475569">睡眠データを記録してください</div>
                    </div>
                    """, unsafe_allow_html=True)

            with r1c2:
                color_map = {
                    "green":  ("#064e3b", "#00ff88"),
                    "orange": ("#431407", "#fb923c"),
                    "yellow": ("#422006", "#fbbf24"),
                    "red":    ("#450a0a", "#f87171"),
                    "gray":   ("#1e293b", "#94a3b8"),
                }
                bg, fg = color_map.get(training["color"], ("#1e293b", "#94a3b8"))
                st.markdown(f"""
                <div class="score-card" style="background:linear-gradient(135deg,{bg} 0%,#1e293b 100%)">
                    <div class="score-label">TODAY'S TRAINING</div>
                    <div style="font-size:36px;font-weight:900;color:{fg};margin:12px 0;line-height:1.2">
                        {training["label"]}
                    </div>
                    <div style="font-size:12px;color:#94a3b8;margin-top:8px;line-height:1.5">
                        {training["advice"]}
                    </div>
                </div>
                """, unsafe_allow_html=True)

            with r1c3:
                est = bio_age.get("estimated_age")
                conf = bio_age.get("confidence", 0)
                offset = bio_age.get("offset", 0)
                if est is not None:
                    diff_str = f"実年齢より {abs(offset):.0f}歳 {'若い' if offset < 0 else '老化'}" if offset != 0 else "実年齢と同等"
                    diff_color = "#00ff88" if offset <= 0 else "#f87171"
                    st.markdown(f"""
                    <div class="bio-age-card">
                        <div class="score-label">BIOLOGICAL AGE</div>
                        <div class="score-number" style="color:white">{est}</div>
                        <div style="font-size:14px;color:{diff_color};font-weight:600;margin-top:4px">
                            {diff_str}
                        </div>
                        <div style="font-size:11px;color:rgba(255,255,255,0.5);margin-top:8px">
                            信頼度 {conf}%
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class="bio-age-card">
                        <div class="score-label">BIOLOGICAL AGE</div>
                        <div class="score-number" style="color:rgba(255,255,255,0.3)">—</div>
                        <div style="font-size:12px;color:rgba(255,255,255,0.4);margin-top:8px">
                            データが蓄積されると表示されます
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Row 2: Recovery components + AI Advice ────────────────────────
            r2c1, r2c2 = st.columns([1, 1.4])

            with r2c1:
                st.markdown("#### 📈 回復スコア内訳")
                comps = recovery.get("components", {})
                if comps:
                    for label, info in comps.items():
                        s = info["score"]
                        bar_color = "#00ff88" if s >= 80 else ("#fbbf24" if s >= 50 else "#f87171")
                        filled = int(s / 10)
                        bar = "█" * filled + "░" * (10 - filled)
                        st.markdown(f"""
                        <div class="comp-row">
                            <span>{label}</span>
                            <span style="font-family:monospace;color:{bar_color}">{bar}</span>
                            <span style="color:{bar_color};font-weight:700">{info['value']}</span>
                        </div>
                        """, unsafe_allow_html=True)

                    # Bio age components
                    bio_comps = bio_age.get("components", [])
                    if bio_comps:
                        st.markdown("<br>**🧬 Biological Age 要因**", unsafe_allow_html=True)
                        for c in bio_comps:
                            impact_color = "#00ff88" if c["impact"] == "若い" else ("#94a3b8" if c["impact"] == "標準" else "#f87171")
                            st.markdown(f"""
                            <div class="comp-row">
                                <span>{c['label']}</span>
                                <span>{c['detail']}</span>
                                <span style="color:{impact_color};font-weight:600">{c['impact']}</span>
                            </div>
                            """, unsafe_allow_html=True)
                else:
                    st.caption("睡眠データがありません")

            with r2c2:
                st.markdown("#### 🤖 今日のAIアドバイス")
                if st.button("AIアドバイスを生成", type="primary", use_container_width=True):
                    with st.spinner("AIが分析中..."):
                        try:
                            prompt = build_ai_prompt(dd)
                            advice = get_ai_advice(prompt, provider, api_key)
                            st.session_state["ai_advice"] = advice
                        except Exception as e:
                            st.error(f"AIエラー: {e}")

                advice = st.session_state.get("ai_advice", {})
                if advice:
                    icons = {
                        "today_priority":    ("🎯", "今日の最優先事項"),
                        "nutrition_advice":  ("🍽️", "栄養アドバイス"),
                        "sleep_target":      ("😴", "今夜の睡眠目標"),
                        "supplement_timing": ("💊", "サプリタイミング"),
                        "training_note":     ("💪", "トレーニングメモ"),
                    }
                    for key, (icon, label) in icons.items():
                        val = advice.get(key)
                        if val:
                            st.markdown(f"""
                            <div class="advice-card">
                                <strong style="color:#a78bfa">{icon} {label}</strong><br>
                                {val}
                            </div>
                            """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div style="color:#475569;font-size:13px;padding:20px;text-align:center">
                        ボタンを押すとAIが今日のデータを分析して<br>
                        パーソナライズされたアドバイスを生成します
                    </div>
                    """, unsafe_allow_html=True)

            # ── Row 3: Recent Sleep trend ─────────────────────────────────────
            if sleep_recs:
                st.markdown("<br>#### 📉 直近14日 睡眠トレンド", unsafe_allow_html=True)
                try:
                    import pandas as pd
                    import plotly.graph_objects as go

                    df_sleep = pd.DataFrame(sleep_recs).sort_values("date")
                    fig = go.Figure()

                    if "total_sleep_h" in df_sleep.columns:
                        fig.add_trace(go.Bar(
                            x=df_sleep["date"].astype(str),
                            y=df_sleep["total_sleep_h"],
                            name="総睡眠",
                            marker_color="#6366f1",
                            opacity=0.8,
                        ))
                    if "deep_sleep_h" in df_sleep.columns:
                        fig.add_trace(go.Bar(
                            x=df_sleep["date"].astype(str),
                            y=df_sleep["deep_sleep_h"],
                            name="深睡眠",
                            marker_color="#00ff88",
                            opacity=0.9,
                        ))
                    # Target line
                    fig.add_hline(y=7.5, line_dash="dot", line_color="#fbbf24",
                                  annotation_text="目標 7.5h", annotation_font_color="#fbbf24")

                    fig.update_layout(
                        barmode="overlay",
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(15,15,30,0.8)",
                        font=dict(color="#94a3b8", size=11),
                        height=220,
                        margin=dict(l=0, r=0, t=10, b=0),
                        legend=dict(orientation="h", y=1.1, x=0,
                                    font=dict(color="#94a3b8")),
                        xaxis=dict(gridcolor="rgba(255,255,255,0.05)"),
                        yaxis=dict(gridcolor="rgba(255,255,255,0.05)", title="時間 (h)"),
                    )
                    st.plotly_chart(fig, use_container_width=True)
                except ImportError:
                    st.caption("グラフ表示にはplotlyが必要です: pip install plotly")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — DATA INPUT
# ═══════════════════════════════════════════════════════════════════════════════
with tab_input:
    import pandas as pd

    col1, col2 = st.columns([1, 1])

    with col1:
        sheet_label = st.selectbox("① シート種別", list(SHEET_OPTIONS.keys()))
        sheet_type = SHEET_OPTIONS[sheet_label]

        selected_date = st.date_input("② 日付", value=datetime.date.today())

        input_mode = st.radio(
            "③ 入力方法",
            ["📁 ファイルを選ぶ", "📷 カメラで撮影"],
            horizontal=True,
        )

        uploaded_files = []

        if input_mode == "📁 ファイルを選ぶ":
            files = st.file_uploader(
                "画像をアップロード（複数可・Nutritionは省略可）",
                type=["png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
            )
            if files:
                uploaded_files = files
        else:
            camera_photo = st.camera_input("カメラで撮影")
            if camera_photo:
                if "camera_photos" not in st.session_state:
                    st.session_state["camera_photos"] = []
                existing_names = [f.name for f in st.session_state["camera_photos"]]
                if camera_photo.name not in existing_names:
                    st.session_state["camera_photos"].append(camera_photo)
            if "camera_photos" in st.session_state and st.session_state["camera_photos"]:
                uploaded_files = st.session_state["camera_photos"]
                st.caption(f"📷 {len(uploaded_files)} 枚撮影済み")
                _cam_cols = st.columns(min(len(uploaded_files), 3))
                for i, f in enumerate(uploaded_files):
                    _cam_cols[i % 3].image(f, use_container_width=True)
                if st.button("🗑️ 撮影リセット", use_container_width=True):
                    del st.session_state["camera_photos"]
                    st.rerun()

        # Performance のみ：テキスト入力欄（任意）
        performance_text = ""
        if sheet_type == "performance":
            st.markdown("---")
            performance_text = st.text_area(
                "直接入力（任意）",
                placeholder="例：Push up: 30, Chin up: 12, Squat: 50\n（スクリーンショットがなければここに数字を入力してください）",
                height=80,
            )

        # Nutrition のみ：食事タイプ選択 + 説明欄 + サプリ
        nutrition_description = ""
        nutrition_meal_type = "dinner"
        nutrition_meal_time = ""
        if sheet_type == "nutrition":
            st.markdown("---")
            nutrition_meal_type = st.selectbox(
                "④ 食事タイプ",
                options=["breakfast", "lunch", "dinner", "snacks"],
                format_func=lambda x: {"breakfast": "🌅 Breakfast（朝食）", "lunch": "☀️ Lunch（昼食）",
                                        "dinner": "🌙 Dinner（夕食）", "snacks": "🍎 Snacks（間食）"}[x],
                index=2,
            )
            nutrition_meal_time = st.text_input(
                "⑤ 食事時間（任意）",
                placeholder="例：午後1時、13:00、8:30",
                help="Breakfast time / Lunch time / Dinner time 行に書き込まれます",
            )
            nutrition_description = st.text_area(
                "⑥ 食事の説明（任意）",
                placeholder="例：チキンカレー（ご飯200g、カレールー150g）、サラダ、水\n食材・量・料理名を書くとAIの精度が上がります",
                height=100,
                help="複数画像の場合、この説明は全画像に共通で使われます",
            )
            st.markdown("---")
            st.caption("💊 サプリメント（今日飲んだものをチェック）")
            from excel_writer.nutrition_writer import SUPPLEMENT_NAMES
            supplement_cols = st.columns(2)
            selected_supplements = []
            for i, name in enumerate(SUPPLEMENT_NAMES):
                if supplement_cols[i % 2].checkbox(name, key=f"supp_{name}"):
                    selected_supplements.append(name)
            if selected_supplements and st.button("💊 サプリを書き込む", use_container_width=True):
                with st.spinner("書き込み中..."):
                    try:
                        if _IS_CLOUD:
                            from gdrive_helper import download_to_temp, upload_from_path
                            _supp_path = download_to_temp()
                        else:
                            _supp_path = excel_path
                        from excel_writer.nutrition_writer import write_supplement_data
                        n = write_supplement_data(selected_supplements, selected_date, _supp_path,
                                                  sheet_prefix=_USER_PREFIX)
                        if _IS_CLOUD:
                            upload_from_path(_supp_path)
                            pathlib.Path(_supp_path).unlink(missing_ok=True)
                            st.session_state.pop("cloud_excel_tmp_path", None)
                        st.success(f"✅ {n} 件のサプリを書き込みました（{selected_date}）")
                    except Exception as e:
                        st.error(f"書き込みエラー: {e}")

        if uploaded_files and input_mode == "📁 ファイルを選ぶ":
            if len(uploaded_files) == 1:
                st.image(uploaded_files[0], caption=uploaded_files[0].name, use_container_width=True)
            else:
                st.caption(f"📎 {len(uploaded_files)} 枚アップロード済み")
                _prev_cols = st.columns(min(len(uploaded_files), 3))
                for i, f in enumerate(uploaded_files):
                    _prev_cols[i % 3].image(f, caption=f.name, use_container_width=True)

    with col2:
        # Nutrition/Performance はテキストのみでも動作可。それ以外は画像必須。
        _can_proceed = bool(uploaded_files)
        if not _can_proceed:
            if sheet_type == "nutrition" and nutrition_description.strip():
                _can_proceed = True
            elif sheet_type == "performance" and performance_text.strip():
                _can_proceed = True

        if not _can_proceed:
            if sheet_type == "nutrition":
                st.info("← 画像をアップロードするか、食事の説明を入力してください")
            elif sheet_type == "performance":
                st.info("← スクリーンショットをアップロードするか、テキスト欄に数値を入力してください")
            else:
                st.info("← スクリーンショットをアップロードしてください（複数可）")
        else:
            if sheet_type == "nutrition":
                extract_btn_label = "⑥ データを抽出する"
            elif sheet_type == "performance":
                extract_btn_label = "④ データを確認する"
            else:
                extract_btn_label = "④ データを抽出する"

            if st.button(extract_btn_label, type="primary", use_container_width=True):
                if not api_key:
                    st.error("APIキーを入力してください（サイドバー）")
                elif _IS_CLOUD:
                    from gdrive_helper import is_configured
                    if not is_configured():
                        st.error("Google Drive未設定です")
                    else:
                        _run_extract = True
                elif not _IS_CLOUD and (not excel_path or not pathlib.Path(excel_path).exists()):
                    st.error(f"Excelファイルが見つかりません: {excel_path}")
                else:
                    _run_extract = True

                if st.session_state.get("_run_extract") or locals().get("_run_extract"):
                    all_results = []
                    if sheet_type in ("nutrition", "performance") and not uploaded_files:
                        spinner_msg = "AIがテキストから値を推定中..."
                        fake_files = [None]
                    else:
                        spinner_msg = f"AIが {len(uploaded_files)} 枚を解析中..."
                        fake_files = uploaded_files

                    with st.spinner(spinner_msg):
                        for uploaded_file in fake_files:
                            if uploaded_file is None:
                                tmp_path = None
                                fname = "（テキストのみ）"
                            else:
                                suffix = pathlib.Path(uploaded_file.name).suffix
                                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                                    tmp.write(uploaded_file.getvalue())
                                    tmp_path = tmp.name
                                fname = uploaded_file.name

                            try:
                                if sheet_type == "nutrition":
                                    from extractors.nutrition_extractor import extract_nutrition_data
                                    data = extract_nutrition_data(tmp_path, provider, api_key,
                                                                  description=nutrition_description)
                                elif sheet_type == "sleep":
                                    from extractors.sleep_extractor import extract_sleep_data
                                    data = extract_sleep_data(tmp_path, provider=provider, api_key=api_key)
                                elif sheet_type == "labs":
                                    from extractors.labs_extractor import extract_labs_data
                                    data = extract_labs_data(tmp_path, provider, api_key)
                                elif sheet_type == "inbody":
                                    from extractors.inbody_extractor import extract_inbody_data
                                    data = extract_inbody_data(tmp_path, provider, api_key)
                                elif sheet_type == "workout":
                                    from extractors.workout_extractor import extract_workout_data
                                    data = extract_workout_data(tmp_path, provider, api_key)
                                elif sheet_type == "performance":
                                    from extractors.performance_extractor import extract_performance_data
                                    data = extract_performance_data(tmp_path, provider, api_key,
                                                                    text=performance_text)
                                else:
                                    data = {}
                                all_results.append({"file": fname, "data": data, "status": "ok"})
                            except Exception as e:
                                all_results.append({"file": fname, "data": None,
                                                    "status": "error", "error": str(e)})
                            finally:
                                if tmp_path:
                                    pathlib.Path(tmp_path).unlink(missing_ok=True)

                    ok = sum(1 for r in all_results if r["status"] == "ok")
                    ng = len(all_results) - ok
                    if ok:
                        st.success(f"解析完了！ {ok} 枚成功" + (f"、{ng} 枚エラー" if ng else ""))
                    else:
                        st.error("すべての画像でエラーが発生しました")
                    for r in all_results:
                        if r["status"] == "error":
                            st.warning(f"❌ {r['file']}: {r['error']}")

                    st.session_state["all_results"] = all_results
                    st.session_state["sheet_type"] = sheet_type
                    st.session_state["selected_date"] = selected_date
                    st.session_state["excel_path"] = excel_path
                    st.session_state["nutrition_description"] = nutrition_description
                    st.session_state["nutrition_meal_type"] = nutrition_meal_type
                    st.session_state["nutrition_meal_time"] = nutrition_meal_time
                    st.session_state["performance_text"] = performance_text

            # ── プレビュー ──────────────────────────────────────────────────────
            if "all_results" in st.session_state:
                all_results = st.session_state["all_results"]
                sheet_type_saved = st.session_state["sheet_type"]
                date_saved = st.session_state["selected_date"]

                if _IS_CLOUD:
                    if "cloud_excel_tmp_path" not in st.session_state:
                        from gdrive_helper import download_to_temp
                        st.session_state["cloud_excel_tmp_path"] = download_to_temp()
                    excel_path_saved = st.session_state["cloud_excel_tmp_path"]
                else:
                    excel_path_saved = excel_path

                meal_type_saved = st.session_state.get("nutrition_meal_type", "dinner")
                meal_time_saved = st.session_state.get("nutrition_meal_time", "")

                if sheet_type_saved != sheet_type:
                    del st.session_state["all_results"]
                    st.info("シートを変更しました。再度抽出してください。")
                else:
                    ok_results = [r for r in all_results if r["status"] == "ok"]
                    if ok_results:
                        def merge_dicts(dicts):
                            merged = {}
                            for d in dicts:
                                if not isinstance(d, dict):
                                    continue
                                for k, v in d.items():
                                    if v is not None:
                                        existing = merged.get(k, {})
                                        if isinstance(v, dict) and isinstance(existing, dict):
                                            merged[k] = merge_dicts([existing, v])
                                        else:
                                            merged[k] = v
                            return merged

                        if sheet_type == "nutrition":
                            display_items = [{"label": r["file"], "data": r["data"]} for r in ok_results]
                        else:
                            merged_data = merge_dicts([r["data"] for r in ok_results])
                            files = ", ".join(r["file"] for r in ok_results)
                            display_items = [{"label": files, "data": merged_data}]
                            if len(ok_results) > 1:
                                st.info(f"📎 {len(ok_results)} 枚をマージして {date_saved} に書き込みます")

                        preview_title = "⑦ 抽出結果プレビュー" if sheet_type == "nutrition" else "⑤ 抽出結果プレビュー"
                        st.subheader(preview_title)

                        for item in display_items:
                            data = item["data"]
                            if len(display_items) > 1:
                                st.markdown(f"**📄 {item['label']}**")

                            if sheet_type == "nutrition":
                                kcal = data.get("calories_kcal")
                                summary = data.get("description_summary", "")
                                m1, m2, m3, m4 = st.columns(4)
                                m1.metric("カロリー", f"{kcal:.0f} kcal" if kcal else "—")
                                m2.metric("タンパク質", f"{data.get('protein_g', 0):.1f} g")
                                m3.metric("脂質", f"{data.get('fat_g', 0):.1f} g")
                                m4.metric("炭水化物", f"{data.get('carbohydrates_g', 0):.1f} g")
                                if summary:
                                    st.caption(f"🍽️ {meal_type_saved.capitalize()}  —  {summary}")

                            with st.expander("生データ（JSON）", expanded=False):
                                st.json(data)

                            try:
                                if sheet_type == "nutrition":
                                    from excel_writer.nutrition_writer import build_nutrition_preview
                                    import openpyxl
                                    from excel_writer.writer_base import ExcelWriter
                                    wb_tmp = openpyxl.load_workbook(excel_path_saved)
                                    ws_tmp = wb_tmp["Nutrition"]
                                    col_idx_tmp = ExcelWriter.find_date_column(ws_tmp, date_saved) or (ws_tmp.max_column + 1)
                                    preview = build_nutrition_preview(data, meal_type_saved, col_idx_tmp)
                                elif sheet_type == "sleep":
                                    from excel_writer.sleep_writer import build_sleep_preview
                                    import openpyxl
                                    from excel_writer.writer_base import ExcelWriter
                                    wb_tmp = openpyxl.load_workbook(excel_path_saved)
                                    ws_tmp = wb_tmp[ExcelWriter.SHEET_SLEEP]
                                    col_idx = ExcelWriter.find_date_column(ws_tmp, date_saved) or (ws_tmp.max_column + 1)
                                    preview = build_sleep_preview(data, col_idx)
                                elif sheet_type == "labs":
                                    from excel_writer.labs_writer import build_labs_preview
                                    import openpyxl
                                    from excel_writer.writer_base import ExcelWriter
                                    wb_tmp = openpyxl.load_workbook(excel_path_saved)
                                    ws_tmp = wb_tmp[ExcelWriter.SHEET_LABS]
                                    col_idx = ExcelWriter.find_date_column(ws_tmp, date_saved) or 2
                                    preview = build_labs_preview(ws_tmp, data.get("results", []), col_idx)
                                elif sheet_type == "inbody":
                                    from excel_writer.inbody_writer import build_inbody_preview
                                    preview = build_inbody_preview(data, row_idx=2)
                                elif sheet_type in ("workout", "performance"):
                                    from excel_writer.workout_writer import build_workout_preview
                                    import openpyxl as _opx_pv
                                    try:
                                        _wb_pv = _opx_pv.load_workbook(excel_path_saved)
                                        _ws_pv = next((wb_[n] for n in _wb_pv.sheetnames if n.strip().lower() == "workout"), None)
                                    except Exception:
                                        _ws_pv = None
                                    preview = build_workout_preview(data, row_idx=2, ws=_ws_pv)
                                else:
                                    preview = []

                                if sheet_type == "labs":
                                    df = pd.DataFrame([{k: v for k, v in row.items() if not k.startswith("_")} for row in preview])
                                    def highlight_confidence(row):
                                        conf = row.get("信頼度", 1.0)
                                        if conf == 0.0: return ["background-color: #ffcccc"] * len(row)
                                        elif conf < 1.0: return ["background-color: #fff3cd"] * len(row)
                                        return [""] * len(row)
                                    st.dataframe(df.style.apply(highlight_confidence, axis=1), use_container_width=True)
                                    unmatched = sum(1 for r in preview if not r["_write"])
                                    if unmatched:
                                        st.warning(f"⚠️ {unmatched} 項目がマッチせず書き込みスキップされます")
                                else:
                                    df = pd.DataFrame([{k: v for k, v in row.items() if not k.startswith("_")} for row in preview])
                                    st.dataframe(df, use_container_width=True)
                            except Exception as e:
                                st.warning(f"プレビュー生成中のエラー: {e}")

                            if len(display_items) > 1:
                                st.divider()

                        # ── 書き込みボタン ────────────────────────────────────
                        n_label = f"（{len(ok_results)} 枚分）" if len(ok_results) > 1 else ""
                        write_btn_label = f"{'⑧' if sheet_type == 'nutrition' else '⑥'} Excelに書き込む ✅{n_label}"

                        if st.button(write_btn_label, type="primary", use_container_width=True):
                            from excel_writer.writer_base import ExcelFileLockError
                            with st.spinner("Excelに書き込み中..."):
                                try:
                                    if _IS_CLOUD:
                                        from gdrive_helper import download_to_temp
                                        excel_path_saved = download_to_temp()

                                    if sheet_type == "nutrition":
                                        from excel_writer.nutrition_writer import write_nutrition_data
                                        merged_nutrition = merge_dicts([r["data"] for r in ok_results])
                                        write_nutrition_data(merged_nutrition, date_saved, meal_type_saved,
                                                             excel_path_saved, meal_time=meal_time_saved,
                                                             sheet_prefix=_USER_PREFIX)
                                        meal_label = {"breakfast": "Breakfast", "lunch": "Lunch",
                                                      "dinner": "Dinner", "snacks": "Snacks"}.get(meal_type_saved, meal_type_saved)
                                        st.success(f"✅ Nutrition [{meal_label}] に書き込みました！（{date_saved}）")
                                    else:
                                        merged = display_items[0]["data"]
                                        if sheet_type == "sleep":
                                            from excel_writer.sleep_writer import write_sleep_data
                                            write_sleep_data(merged, date_saved, excel_path_saved,
                                                             sheet_prefix=_USER_PREFIX)
                                        elif sheet_type == "labs":
                                            from excel_writer.labs_writer import write_labs_data
                                            write_labs_data(merged.get("results", []), date_saved, excel_path_saved,
                                                            sheet_prefix=_USER_PREFIX)
                                        elif sheet_type == "inbody":
                                            from excel_writer.inbody_writer import write_inbody_data
                                            write_inbody_data(merged, date_saved, excel_path_saved,
                                                              sheet_prefix=_USER_PREFIX)
                                        elif sheet_type in ("workout", "performance"):
                                            from excel_writer.workout_writer import write_workout_data, get_column_a_dump
                                            try:
                                                col_a = get_column_a_dump(excel_path_saved,
                                                                          sheet_prefix=_USER_PREFIX)
                                                if col_a:
                                                    with st.expander("🔍 Workout sheet 行構造", expanded=False):
                                                        st.dataframe(pd.DataFrame(col_a, columns=["行", "列A", "列B"]),
                                                                     use_container_width=True)
                                            except Exception:
                                                pass
                                            for r in ok_results:
                                                dbg = write_workout_data(r["data"], date_saved, excel_path_saved,
                                                                         sheet_prefix=_USER_PREFIX)
                                                st.caption(f"✍️ {dbg.get('workout_type')} → col {dbg.get('col')} | "
                                                           f"{list(dbg.get('written', {}).keys())}")
                                        st.success(f"✅ Workout シートに書き込みました！（{date_saved}）")

                                    if _IS_CLOUD:
                                        from gdrive_helper import upload_from_path
                                        upload_from_path(excel_path_saved)
                                        pathlib.Path(excel_path_saved).unlink(missing_ok=True)
                                        st.session_state.pop("cloud_excel_tmp_path", None)
                                        st.session_state.pop("dashboard_data", None)
                                        st.session_state.pop("dashboard_excel_path", None)
                                        st.success("☁️ Google Driveに自動保存しました")

                                    del st.session_state["all_results"]

                                except ExcelFileLockError as e:
                                    st.error(str(e))
                                except Exception as e:
                                    st.error(f"書き込みエラー: {e}")
                                    with st.expander("詳細"):
                                        import traceback
                                        st.code(traceback.format_exc())
