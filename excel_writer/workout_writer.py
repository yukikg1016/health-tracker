"""
Workout writer — label-based row finding.
Scans column A/B of the Workout sheet to find section headers and field rows,
rather than relying on hardcoded row numbers that may differ between spreadsheets.
"""
import datetime
import openpyxl
from .writer_base import ExcelWriter

SHEET_NAME = "Workout"

# Search terms for section headers (checked against column A content, case-insensitive)
SECTION_SEARCH = {
    "core training":      ["core training", "core", "コアトレ"],
    "strength training":  ["strength training", "strength", "ストレングス"],
    "outdoor run":        ["outdoor run", "run", "ランニング", "アウトドアラン"],
    "outdoor walk":       ["outdoor walk", "walk", "ウォーキング", "アウトドアウォーク"],
    "activity rings":     ["activity rings", "activity", "アクティビティ"],
    "performance":        ["performance", "体力測定", "パフォーマンス"],
}

# Search terms for field labels (checked within a section, case-insensitive)
FIELD_SEARCH = {
    "workout_time_min":    ["workout time", "ワークアウト時間", "time"],
    "elapsed_time_min":    ["elapsed", "経過時間"],
    "distance_km":         ["distance", "距離"],
    "active_kcal":         ["active kcal", "active cal", "アクティブ"],
    "total_kcal":          ["total kcal", "合計"],
    "avg_heart_rate_bpm":  ["heart rate", "心拍数", "bpm", "avg heart"],
    "avg_pace_min_per_km": ["pace", "ペース"],
    "avg_power_watts":     ["power", "パワー"],
    "avg_cadence_spm":     ["cadence", "ケイデンス"],
    "effort":              ["effort", "難易度"],
    "move_kcal":           ["move", "ムーブ"],
    "exercise_min":        ["exercise", "エクササイズ"],
    "stand_hours":         ["stand", "スタンド"],
    "step_count":          ["step count", "歩数"],
    "step_distance_km":    ["step distance", "歩行距離"],
    # Performance
    "push_up":             ["push up", "pushup", "腕立て"],
    "chin_up":             ["chin up", "chinup", "懸垂", "チンアップ"],
    "squat":               ["squat", "スクワット"],
}

# Fields belonging to each workout type
WORKOUT_FIELDS = {
    "core training":      ["workout_time_min", "elapsed_time_min", "active_kcal", "total_kcal",
                           "avg_heart_rate_bpm", "effort"],
    "strength training":  ["workout_time_min", "elapsed_time_min", "active_kcal", "total_kcal",
                           "avg_heart_rate_bpm", "effort"],
    "outdoor run":        ["workout_time_min", "distance_km", "active_kcal", "total_kcal",
                           "avg_power_watts", "avg_cadence_spm", "avg_pace_min_per_km",
                           "avg_heart_rate_bpm", "effort"],
    "outdoor walk":       ["workout_time_min", "distance_km", "active_kcal", "total_kcal",
                           "avg_pace_min_per_km", "avg_heart_rate_bpm", "effort"],
    "activity rings":     ["move_kcal", "exercise_min", "stand_hours", "step_count", "step_distance_km"],
    "performance":        ["push_up", "chin_up", "squat"],
}

# Hardcoded fallback row maps (used if label search fails)
_FALLBACK_ROWS = {
    "core training":     {"workout_time_min": 4, "elapsed_time_min": 5, "active_kcal": 6,
                          "total_kcal": 7, "avg_heart_rate_bpm": 8, "effort": 9},
    "strength training": {"workout_time_min": 4, "elapsed_time_min": 5, "active_kcal": 6,
                          "total_kcal": 7, "avg_heart_rate_bpm": 8, "effort": 9},
    "outdoor run":       {"workout_time_min": 13, "distance_km": 14, "active_kcal": 15,
                          "total_kcal": 16, "avg_power_watts": 17, "avg_cadence_spm": 18,
                          "avg_pace_min_per_km": 19, "avg_heart_rate_bpm": 20, "effort": 21},
    "outdoor walk":      {"workout_time_min": 25, "distance_km": 26, "active_kcal": 27,
                          "total_kcal": 28, "avg_pace_min_per_km": 29, "avg_heart_rate_bpm": 30,
                          "effort": 31},
    "activity rings":    {"move_kcal": 36, "exercise_min": 37, "stand_hours": 38,
                          "step_count": 39, "step_distance_km": 40},
    "performance":       {"push_up": 45, "chin_up": 46, "squat": 47},
}


def _cell_text(ws, row, col) -> str:
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def _matches(text: str, terms: list) -> bool:
    text_l = text.lower()
    return any(t.lower() in text_l for t in terms)


def scan_workout_sheet(ws) -> dict:
    """
    Scan the worksheet and return a mapping:
      { section_key: { field_key: row_number } }
    where section_key matches SECTION_SEARCH keys, field_key matches FIELD_SEARCH keys.
    """
    max_row = ws.max_row

    # Step 1: find section header rows
    section_rows = {}  # section_key → row number where header found
    for row in range(1, max_row + 1):
        for col in [1, 2]:
            text = _cell_text(ws, row, col)
            if not text:
                continue
            for sec_key, terms in SECTION_SEARCH.items():
                if sec_key not in section_rows and _matches(text, terms):
                    section_rows[sec_key] = row

    # Step 2: for each section, scan the next N rows for field labels
    result = {}
    section_items = sorted(section_rows.items(), key=lambda x: x[1])

    for i, (sec_key, sec_row) in enumerate(section_items):
        # End of this section = start of next section (or max_row)
        if i + 1 < len(section_items):
            end_row = section_items[i + 1][1] - 1
        else:
            end_row = min(sec_row + 30, max_row)

        fields_found = {}
        expected_fields = WORKOUT_FIELDS.get(sec_key, [])
        for field in expected_fields:
            terms = FIELD_SEARCH.get(field, [field])
            for row in range(sec_row + 1, end_row + 1):
                for col in [1, 2]:
                    text = _cell_text(ws, row, col)
                    if text and _matches(text, terms):
                        fields_found[field] = row
                        break
                if field in fields_found:
                    break
        result[sec_key] = fields_found

    return result


def _get_row_map(workout_type: str, ws=None) -> dict:
    """Return field→row map. Uses label-based scan if ws provided, else fallback."""
    if not workout_type:
        return {}
    key = workout_type.lower().strip()
    # Normalize key
    for sec_key in SECTION_SEARCH:
        if key == sec_key or _matches(key, SECTION_SEARCH[sec_key]):
            key = sec_key
            break

    if ws is not None:
        scan = scan_workout_sheet(ws)
        if key in scan and scan[key]:
            return scan[key]

    return _FALLBACK_ROWS.get(key, _FALLBACK_ROWS.get("core training", {}))


def write_workout_data(data: dict, target_date, excel_path: str,
                       sheet_prefix: str = "") -> dict:
    """Write workout data. Returns debug info dict."""
    if isinstance(target_date, str):
        target_date = datetime.date.fromisoformat(target_date)

    writer = ExcelWriter(excel_path)
    wb = writer.load()

    # Find sheet (flexible name matching)
    ws = None
    target_name = (sheet_prefix + SHEET_NAME).strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target_name:
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(f"Workoutシートが見つかりません。シート一覧: {wb.sheetnames}")

    # Find date column
    col_idx = ExcelWriter.find_date_column(ws, target_date)
    if col_idx is None:
        col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx).value = target_date.strftime("%Y.%m.%d")

    workout_type = data.get("workout_type", "")
    row_map = _get_row_map(workout_type, ws)

    debug = {"sheet": ws.title, "col": col_idx, "workout_type": workout_type,
             "row_map": row_map, "written": {}}

    for field, row in row_map.items():
        val = data.get(field)
        if val is not None:
            ws.cell(row=row, column=col_idx).value = val
            debug["written"][field] = {"row": row, "val": val}

    # Activity Rings: write "current/goal" format
    if workout_type and "activity" in workout_type.lower():
        move_row = row_map.get("move_kcal")
        ex_row = row_map.get("exercise_min")
        st_row = row_map.get("stand_hours")

        move = data.get("move_kcal")
        move_goal = data.get("move_goal_kcal")
        if move_row and move is not None and move_goal is not None:
            ws.cell(row=move_row, column=col_idx).value = f"{move}/{move_goal}"
        elif move_row and move is not None:
            ws.cell(row=move_row, column=col_idx).value = move

        ex = data.get("exercise_min")
        ex_goal = data.get("exercise_goal_min")
        if ex_row and ex is not None and ex_goal is not None:
            ws.cell(row=ex_row, column=col_idx).value = f"{ex}/{ex_goal}"
        elif ex_row and ex is not None:
            ws.cell(row=ex_row, column=col_idx).value = ex

        st = data.get("stand_hours")
        st_goal = data.get("stand_goal_hours")
        if st_row and st is not None and st_goal is not None:
            ws.cell(row=st_row, column=col_idx).value = f"{st}/{st_goal}"
        elif st_row and st is not None:
            ws.cell(row=st_row, column=col_idx).value = st

    writer.save(wb)
    return debug


def get_column_a_dump(excel_path: str, sheet_prefix: str = "") -> list:
    """Return list of (row, col_a_value, col_b_value) for debugging."""
    wb = openpyxl.load_workbook(excel_path)
    ws = None
    target_name = (sheet_prefix + SHEET_NAME).strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target_name:
            ws = wb[name]
            break
    if ws is None:
        return []
    result = []
    for row in range(1, min(ws.max_row + 1, 60)):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        if a is not None or b is not None:
            result.append((row, a, b))
    return result


def build_workout_preview(data: dict, row_idx: int, ws=None) -> list:
    workout_type = data.get("workout_type", "")
    row_map = _get_row_map(workout_type, ws)

    label_names = {
        "workout_time_min":    "Workout Time (min)",
        "elapsed_time_min":    "Elapsed Time (min)",
        "distance_km":         "Distance (km)",
        "active_kcal":         "Active kcal",
        "total_kcal":          "Total kcal",
        "avg_heart_rate_bpm":  "Avg Heart Rate (bpm)",
        "avg_pace_min_per_km": "Avg Pace (min/km)",
        "avg_power_watts":     "Avg Power (W)",
        "avg_cadence_spm":     "Avg Cadence (spm)",
        "effort":              "Effort",
        "move_kcal":           "Move (kcal)",
        "move_goal_kcal":      "Move Goal (kcal)",
        "exercise_min":        "Exercise (min)",
        "exercise_goal_min":   "Exercise Goal (min)",
        "stand_hours":         "Stand (h)",
        "stand_goal_hours":    "Stand Goal (h)",
        "step_count":          "Step Count",
        "step_distance_km":    "Step Distance (km)",
        "push_up":             "Push Up (reps)",
        "chin_up":             "Chin Up (reps)",
        "squat":               "Squat (reps)",
    }

    rows = [{"項目": "Type", "値": workout_type, "行": "-"}]
    for field, row in row_map.items():
        val = data.get(field)
        if val is not None:
            rows.append({"項目": label_names.get(field, field), "値": val, "行": row})

    # Activity Rings extras
    if workout_type and "activity" in workout_type.lower():
        for field in ["move_goal_kcal", "exercise_goal_min", "stand_goal_hours"]:
            val = data.get(field)
            if val is not None and not any(r["項目"] == label_names.get(field) for r in rows):
                rows.append({"項目": label_names.get(field, field), "値": val, "行": "-"})

    return rows
