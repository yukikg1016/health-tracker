from __future__ import annotations

"""Excel書き込み基底クラス"""

import datetime
import subprocess
from pathlib import Path

import openpyxl


class ExcelFileLockError(Exception):
    pass


class ExcelWriter:
    # Excelシート名（末尾スペース・全角スペース含む実際の名前）
    SHEET_SLEEP   = "Sleep "
    SHEET_LABS    = "Labs "
    SHEET_INBODY  = "In Body"
    SHEET_MORNING = "朝\u3000☀️"
    SHEET_NIGHT   = "夜\u3000💤"

    def __init__(self, excel_path: str):
        self.excel_path = str(excel_path)

    def load(self) -> openpyxl.Workbook:
        return openpyxl.load_workbook(self.excel_path)

    def save(self, wb: openpyxl.Workbook) -> None:
        try:
            wb.save(self.excel_path)
        except (PermissionError, OSError) as e:
            raise ExcelFileLockError(
                f"Excelファイルを保存できませんでした。Numbers/Excelで開いている場合は閉じてください。\n詳細: {e}"
            )
        # 保存後、開いているExcelに再読み込みを指示する
        self._refresh_excel()

    def _refresh_excel(self) -> None:
        """開いているMicrosoft ExcelでYukiデータのワークブックを閉じてから再度openする。"""
        abs_path = str(Path(self.excel_path).resolve())
        # Step1: Yuki データを含むワークブックを保存せずに閉じる
        close_script = '''
tell application "Microsoft Excel"
    set wbList to every workbook
    repeat with wb in wbList
        try
            if name of wb contains "Yuki" then
                close wb saving false
            end if
        end try
    end repeat
end tell
'''
        try:
            subprocess.run(["osascript", "-e", close_script], timeout=5, capture_output=True)
        except Exception:
            pass
        # Step2: ファイルを開く
        try:
            subprocess.run(["open", abs_path], timeout=5, capture_output=True)
        except Exception:
            pass

    def save_backup(self, wb: openpyxl.Workbook) -> str:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base = Path(self.excel_path)
        backup_path = str(base.parent / f"{base.stem}_backup_{ts}{base.suffix}")
        wb.save(backup_path)
        return backup_path

    @staticmethod
    def find_date_column(ws, target_date: datetime.date) -> int | None:
        """Row 1 を左から右に走査し、target_date と一致する列番号（1-based）を返す。
        datetime, date, 文字列 ('2026.04.21' / '2026/04/21') の各形式に対応。
        """
        for col in range(2, ws.max_column + 2):
            cell_val = ws.cell(row=1, column=col).value
            if cell_val is None:
                continue
            if isinstance(cell_val, datetime.datetime):
                cell_val = cell_val.date()
            if isinstance(cell_val, datetime.date) and cell_val == target_date:
                return col
            # 文字列形式 ('2026.04.21' or '2026/04/21')
            if isinstance(cell_val, str):
                normalized = cell_val.replace("/", ".").strip()
                target_str = target_date.strftime("%Y.%m.%d")
                if normalized == target_str:
                    return col
        return None

    @staticmethod
    def get_or_create_date_column(ws, target_date: datetime.date) -> int:
        """日付列を探し、なければ最終列の右に新しく作成して返す。"""
        col = ExcelWriter.find_date_column(ws, target_date)
        if col is not None:
            return col
        # 新規列を末尾に追加
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = datetime.datetime(
            target_date.year, target_date.month, target_date.day
        )
        return new_col
