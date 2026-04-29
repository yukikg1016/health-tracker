"""
Dashboard data reader — reads recent data from Excel for the dashboard tab.
Scans Sleep, Workout, Nutrition, InBody, Labs sheets.
"""
from __future__ import annotations
import datetime
import openpyxl


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v


def _find_date_col(ws, target_date: datetime.date, row=1) -> int | None:
    """Find column index for a given date in row 1."""
    target_str = target_date.strftime("%Y.%m.%d")
    for col in range(2, ws.max_column + 2):
        v = ws.cell(row=row, column=col).value
        if v is None:
            continue
        vs = str(v).strip()
        if vs == target_str:
            return col
        # Also match datetime objects
        if hasattr(v, "strftime") and v.strftime("%Y.%m.%d") == target_str:
            return col
    return None


def _find_label_row(ws, labels: list[str], col=1) -> int | None:
    """Find row where column A matches any of the given labels (case-insensitive)."""
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if v is None:
            continue
        vs = str(v).strip().lower()
        for label in labels:
            if label.lower() in vs:
                return row
    return None


def _get_recent_dates(ws, n=14) -> list[datetime.date]:
    """Get the most recent n dates from row 1 of a sheet."""
    dates = []
    for col in range(2, ws.max_column + 2):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        if isinstance(v, datetime.datetime):
            dates.append(v.date())
        elif isinstance(v, datetime.date):
            dates.append(v)
        elif isinstance(v, str):
            try:
                dates.append(datetime.datetime.strptime(v.strip(), "%Y.%m.%d").date())
            except Exception:
                pass
    dates.sort(reverse=True)
    return dates[:n]


def read_dashboard_data(excel_path: str, days: int = 14, sheet_prefix: str = "") -> dict:
    """
    Read recent data from all sheets.
    Returns a dict with sleep, workout, nutrition, inbody, labs data
    for the past `days` days.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    # sheet_prefix付きシート名と無しの両方でマップを構築
    p = sheet_prefix.strip().lower()
    sheet_map = {}
    for n in wb.sheetnames:
        key = n.strip().lower()
        sheet_map[key] = wb[n]
        # プレフィックスを除いたキーでも引けるようにする
        if p and key.startswith(p):
            sheet_map[key[len(p):].strip()] = wb[n]

    today = datetime.date.today()
    date_range = [today - datetime.timedelta(days=i) for i in range(days)]

    result = {
        "sleep":     _read_sleep(sheet_map, date_range),
        "workout":   _read_workout(sheet_map, date_range),
        "nutrition": _read_nutrition(sheet_map, date_range),
        "inbody":    _read_inbody(sheet_map),
        "labs":      _read_labs(sheet_map),
    }
    return result


# ── Sleep ─────────────────────────────────────────────────────────────────────
_SLEEP_LABELS = {
    "total_sleep_h":   ["total sleep", "総睡眠", "sleep duration", "☾"],
    "deep_sleep_h":    ["deep sleep", "深い睡眠", "深睡眠", "②"],
    "quality_score":   ["quality", "sleep score", "評価", "★"],
    "resting_hr":      ["resting heart", "安静時心拍", "bpm"],
    "hrv":             ["hrv", "heart rate variability"],
    "sleep_bank":      ["sleep bank", "睡眠バンク"],
}

def _read_sleep(sheet_map: dict, date_range: list) -> list[dict]:
    ws = sheet_map.get("sleep")
    if ws is None:
        return []

    # Find label rows
    label_rows = {}
    for key, labels in _SLEEP_LABELS.items():
        row = _find_label_row(ws, labels)
        if row:
            label_rows[key] = row

    records = []
    for d in date_range:
        col = _find_date_col(ws, d)
        if col is None:
            continue
        rec = {"date": d}
        for key, row in label_rows.items():
            v = _cell(ws, row, col)
            if v is not None:
                try:
                    rec[key] = float(v)
                except (TypeError, ValueError):
                    rec[key] = v
        if len(rec) > 1:
            records.append(rec)
    return records


# ── Workout ───────────────────────────────────────────────────────────────────
_WORKOUT_EFFORT_LABELS = ["effort", "難易度"]
_WORKOUT_TIME_LABELS   = ["workout time", "ワークアウト時間"]
_WORKOUT_ACTIVE_LABELS = ["active kcal", "active cal", "アクティブ"]

def _read_workout(sheet_map: dict, date_range: list) -> list[dict]:
    ws = sheet_map.get("workout")
    if ws is None:
        return []

    effort_row = _find_label_row(ws, _WORKOUT_EFFORT_LABELS)
    time_row   = _find_label_row(ws, _WORKOUT_TIME_LABELS)
    kcal_row   = _find_label_row(ws, _WORKOUT_ACTIVE_LABELS)

    records = []
    for d in date_range:
        col = _find_date_col(ws, d)
        if col is None:
            continue
        rec = {"date": d}
        if effort_row:
            v = _cell(ws, effort_row, col)
            if v is not None:
                try: rec["effort"] = float(v)
                except: pass
        if time_row:
            v = _cell(ws, time_row, col)
            if v is not None:
                try: rec["workout_time_min"] = float(v)
                except: pass
        if kcal_row:
            v = _cell(ws, kcal_row, col)
            if v is not None:
                try: rec["active_kcal"] = float(v)
                except: pass
        if len(rec) > 1:
            records.append(rec)
    return records


# ── Nutrition ─────────────────────────────────────────────────────────────────
_NUTRITION_CALORIE_ROW  = 6   # Total calories
_NUTRITION_PROTEIN_ROW  = 11
_NUTRITION_CARBS_ROW    = 13
_NUTRITION_FAT_ROW      = 15

def _read_nutrition(sheet_map: dict, date_range: list) -> list[dict]:
    ws = sheet_map.get("nutrition")
    if ws is None:
        return []

    records = []
    for d in date_range:
        col = _find_date_col(ws, d)
        if col is None:
            continue
        rec = {"date": d}
        for key, row in [("calories", _NUTRITION_CALORIE_ROW),
                         ("protein_g", _NUTRITION_PROTEIN_ROW),
                         ("carbs_g", _NUTRITION_CARBS_ROW),
                         ("fat_g", _NUTRITION_FAT_ROW)]:
            v = _cell(ws, row, col)
            if v is not None:
                try: rec[key] = float(v)
                except: pass
        if len(rec) > 1:
            records.append(rec)
    return records


# ── InBody ────────────────────────────────────────────────────────────────────
_INBODY_LABELS = {
    "body_fat_pct":   ["body fat", "体脂肪率", "fat %"],
    "muscle_kg":      ["muscle mass", "骨格筋", "筋肉量"],
    "weight_kg":      ["weight", "体重"],
    "bmi":            ["bmi"],
}

def _read_inbody(sheet_map: dict) -> dict | None:
    ws = sheet_map.get("inbody") or sheet_map.get("in body")
    if ws is None:
        return None

    # InBody is column-per-date; find most recent
    dates = _get_recent_dates(ws, n=5)
    if not dates:
        return None

    latest = dates[0]
    col = _find_date_col(ws, latest)
    if col is None:
        return None

    rec = {"date": latest}
    for key, labels in _INBODY_LABELS.items():
        row = _find_label_row(ws, labels)
        if row:
            v = _cell(ws, row, col)
            if v is not None:
                try: rec[key] = float(v)
                except: pass
    return rec if len(rec) > 1 else None


# ── Labs ──────────────────────────────────────────────────────────────────────
_LABS_LABELS = {
    "ferritin":    ["ferritin", "フェリチン"],
    "hemoglobin":  ["hemoglobin", "ヘモグロビン", "hgb"],
    "vitamin_d":   ["vitamin d", "ビタミンd", "25-oh"],
    "testosterone":["testosterone", "テストステロン"],
    "cortisol":    ["cortisol", "コルチゾール"],
}

def _read_labs(sheet_map: dict) -> dict | None:
    ws = sheet_map.get("labs")
    if ws is None:
        return None

    dates = _get_recent_dates(ws, n=5)
    if not dates:
        return None

    latest = dates[0]
    col = _find_date_col(ws, latest)
    if col is None:
        # Labs may use row=date structure; try column 2
        col = 2

    rec = {"date": latest}
    for key, labels in _LABS_LABELS.items():
        row = _find_label_row(ws, labels)
        if row:
            v = _cell(ws, row, col)
            if v is not None:
                try: rec[key] = float(v)
                except: pass
    return rec if len(rec) > 1 else None
